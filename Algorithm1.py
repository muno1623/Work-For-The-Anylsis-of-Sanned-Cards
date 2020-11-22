import cv2
import re
import datetime
from geopy.geocoders import Nominatim
from spellchecker import SpellChecker
from PIL import Image, ImageEnhance
import numpy as np
import pytesseract
from scipy import ndimage, misc
from PIL import Image
from PIL import ImageEnhance, ImageFilter
import image
import pandas as pd
from pandas import ExcelWriter
from openpyxl import load_workbook
from geopy.extra.rate_limiter import RateLimiter

##Creating Data Frame and organizing data ##
data = ['Klasse', 'Name', 'Vorname', 'Akt. Zeichen', 'Beruf', 'Anschrift', 'Bezeichnung', 'Angemeldet am', 'Ende', 'ValidAddress']
df = pd.DataFrame(columns=data)
pd.set_option('display.max_columns', None)
pd.set_option('display.width', None)


pytesseract.pytesseract.tesseract_cmd = 'C:/Program Files (x86)/Tesseract-OCR4.0/tesseract.exe'
path = "C:/Users/Tayab/PycharmProjects/MediaProject/Card1.png"
# Reading Image
img = cv2.imread(path)

#........PRE-PROCESSING.......

### Median Filter ###
medianFilt = ndimage.median_filter(img, size=4)

### Converting image to grey scale ###
grayImage = cv2.cvtColor(medianFilt, cv2.COLOR_BGR2GRAY)

### Median Filter ###
medianFilt = ndimage.median_filter(grayImage, size=4)

### Rescalling image ###
Processed = cv2.resize(medianFilt,(1023,678))
OCR = cv2.resize(medianFilt,(1023,678))


#.......TEXT-EXTRACTION.......
custom_config = r'-c tessedit_char_whitelist=abcdefghijklmnopqrstuvwxyz --psm 6'
custom_oem_psm_config = "--oem 3 --psm 6"
text = pytesseract.image_to_string(OCR, config=custom_config, lang='deu')


# Confidences/Accuracy Percentage
dataFrame = pytesseract.image_to_data(OCR, lang='deu', output_type='data.frame')
conf = dataFrame[['conf', 'text']]
conf = conf[conf.conf != -1]
#lines = conf.groupby('block_num')['text'].apply(list)
print(dataFrame)
print(conf)

### Drawing the bounding boxes on the image ###
h, w = OCR.shape
boxes = pytesseract.image_to_boxes(OCR, config=custom_oem_psm_config, lang='deu') # also include any config options you use

for b in boxes.splitlines():
    b = b.split(' ')
    box = cv2.rectangle(OCR, (int(b[1]), h - int(b[2])), (int(b[3]), h - int(b[4])), (0, 255, 0), 2)



#.......EXTRACTING WORDS FROM STRING.......
### using split() ###
res = text.translate({ord(i): None for i in '$°+'})
res = text.split()
date = re.findall("([0-9]{2}\.[0-9]{1}\.[0-9]{4})", str(res))
classification = re.findall("([0-9]{1}\=[0-9]{2})", str(res))
city = re.findall("Drosden", str(res))
##date = datetime.datetime.strptime(match.group(), '%Y.%m.%d').date()
print(date)
print(classification)
print(city)




#.......OUTPUT.........
### Printing Results ###
#print ("The list of words is : " +  str(res))

### Saving data to file ####
file = open("output.txt","w")
file.write(text)
file.close()


#.......Displaying Image.........

### Original Image ###
#cv2.imshow('Original Image', OCR)

### Pre-Processed Image ###
#cv2.imshow('Pre-Processed Image', Processed)

### Detected Characters in Image ###
cv2.imshow('Detected Characters', box)

nameStr = ''
klasseStr = ''
vornameStr = ''
aktStr = ''
berufStr = ''
ansStr = ''
bezeichnungStr = ''
angemeldetStr = ''
aendeStr = ''
##For Name##
for x in range(len(res)):
    if res[x] == 'Name':
        for y in range(x+1, len(res)):
            if res[y] == 'Klasse':
                break
            nameStr += res[y] + ' '
##For Klasse ##
for x in range(len(res)):
    if res[x] == 'Klasse':
        for y in range(x+1, len(res)):
            if res[y] == 'Vorname':
                break
            klasseStr = res[0] + ' '
##For Vorname ##
for x in range(len(res)):
    if res[x] == 'Vorname':
        for y in range(x+1, len(res)):
            if res[y] == 'Akt.':
                break
            vornameStr += res[y] + ' '
##For Akt. Zeichen ##
for x in range(len(res)):
    if res[x] == 'Akt.':
        for y in range(x+2, len(res)):
            if res[y] == 'Beruf':
                break
            aktStr += res[y] + ' '
##For Beruf ##
for x in range(len(res)):
    if res[x] == 'Beruf':
        for y in range(x+1, len(res)):
            if res[y] == 'Anschrift':
                break
            berufStr += res[y] + ' '
##For Anschrift ##
for x in range(len(res)):
    if res[x] == 'Anschrift':
        for y in range(x+1, len(res)):
            if res[y] == 'Bezeichnung':
                break
            ansStr += res[y] + ' '
##For Bezeichnung ##
for x in range(len(res)):
    if res[x] == 'Bezeichnung':
        for y in range(x+1, len(res)):
            if res[y] == 'Angemeldet':
                break
            bezeichnungStr += res[y] + ' '
##For  Angemeldet ##
for x in range(len(res)):
    if res[x] == 'Angemeldet':
        for y in range(x+1, len(res)):
            if res[y] == 'Ende':
                break
            angemeldetStr += res[y] + ' '
##For  Angemeldet ##
for x in range(len(res)):
    if res[x] == 'Ende':
        for y in range(x+1, len(res)):
            if res[y] == 'Verlängert':
                break
            aendeStr += res[y] + ' '
        else:
            continue
        break

df = df.append({'Klasse': klasseStr,
                'Name': nameStr,
                'Vorname': vornameStr,
                'Akt. Zeichen': aktStr,
                'Beruf': berufStr,
                'Anschrift': 'Max Planck Ring 4',
                'Bezeichnung': bezeichnungStr,
                'Angemeldet am': angemeldetStr,
                'Ende': aendeStr,
                'ValidAddress': ''}, ignore_index=True)
##Validated Address##
geolocator = Nominatim(user_agent="abc")
geocode = RateLimiter(geolocator.geocode, min_delay_seconds=1)
df['ValidAddress'] = df['Anschrift'].apply(geocode)
df.to_csv('OrganizedDataCSV.csv')

writer = pd.ExcelWriter('Excel.xlsx', engine='openpyxl')
# try to open an existing workbook
writer.book = load_workbook('Excel.xlsx')
# copy existing sheets
writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
# read existing file
reader = pd.read_excel(r'Excel.xlsx')
sheet = writer.book.active
if sheet.cell(1, 2).value is None:
    df.to_excel(writer, index = False)
else:
    df.to_excel(writer, index = False,header = False, startrow=len(reader) + 1)
writer.close()

print(df)
cv2.waitKey(0)
cv2.destroyAllWindows()
